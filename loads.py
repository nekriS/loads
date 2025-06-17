
import pandas as pd
from pandas.io.formats import excel
excel.ExcelFormatter.header_style = None
import openpyxl
print(openpyxl.__version__)

def set_column_autowidth(ws, columns, reserve=1.2):
    """
    Устанавливает оптимальную ширину столбцов на основе содержимого.
    """
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Получаем букву столбца (A, B, C, ...)
        
        if column in columns:
            # Находим максимальную длину текста в столбце
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
            # Устанавливаем ширину столбца с небольшим запасом
            adjusted_width = (max_length + 2) * reserve  # Можно изменить коэффициент для более комфортного отображения
            ws.column_dimensions[column].width = adjusted_width

def is_float(element: any) -> bool:
    if element is None: 
        return False
    try:
        float(element)
        return True
    except ValueError:
        return False
    
def getValue(current_str):
    current_str = str(current_str)
    current_str = current_str.lower()
    current_str = current_str.replace(' ','')
    current_str = current_str.replace('a','')
    #print(current_str)
    try:
        if 'm' in current_str:
            return float(current_str[:-1]) * pow(10, -3)
        elif 'u' in current_str:
            return float(current_str[:-1]) * pow(10, -6)
        elif 'n' in current_str:
            return float(current_str[:-1]) * pow(10, -9)
        elif 'p' in current_str:
            return float(current_str[:-1]) * pow(10, -12)
        elif is_float(current_str):
            return float(current_str)
        else:
            return ""
    except:
        return ""

#table = pd.read_csv("loads/report_Loads_2025_06_10_17_48_59.csv")
table = pd.read_csv("loads/report_Loads_2025_06_17_11_47_31.csv")
table = table.fillna('')
col = table.columns



#table = table.drop_duplicates(subset=table.columns[1:])
table['Count'] = int
#print(table)
for index, row in table.iterrows():
    table.at[index,'CURRENT'] = "" if getValue(table.at[index,'CURRENT']) == 0 else getValue(table.at[index,'CURRENT'])
#print(table)

for index, row in table.iterrows():
    count = 1
    for index2, row2 in table.iterrows():
        if index2 > index:
            if (row.drop(col[0]) == row2.drop(col[0])).all():
                count += 1
                table.at[index, col[0]] = table.at[index, col[0]] +", "+ table.at[index2, col[0]]
                
                #row[col[0]] += row2[col[0]]
                table = table.drop(index=index2)
    table.at[index, 'Count'] = count

    
    
table = table.dropna()
#print(table)

voltages = []

#table_new = pd.DataFrame()
tn = []

for index, row in table.iterrows():
    line = row[col[6]]
    lines = line.split(" ")
    for l in lines:
        if l != "" and "+" in l:
            
            if l.find("_") == -1:
                l_short = l
            else:
                l_short = l[0:l.find("_")]
            if is_float(l_short.replace("V", ".")):
                voltageDigit = float(l_short.replace("V", "."))
            else:
                voltageDigit = -1
            if not([l, voltageDigit] in voltages):
                voltages.append([l, voltageDigit])
            
            temp_row = row
            temp_row[col[6]] = l

            
            tn.append(temp_row.values.tolist())

            
            #table_new = pd.concat([table_new, temp_row], ignore_index=True)
            
table_new = pd.DataFrame(tn)
table_new = table_new.drop_duplicates()
#print(table_new)



table_new = table_new.drop([2, 4], axis=1)
table_new[8] = ""
table_new[9] = ""
table_new = table_new.reindex(columns=[0, 1, 6, 3, 5, 8, 7, 9])


voltages_table = pd.DataFrame(voltages)
voltages_table = voltages_table.sort_values(1, ascending=False)



header = ["Обозначение", "Потребитель", "NetName", "Ток (CIP)", "Ток (Схема)", "Ток (Вручную)", "Количество", "Итоговый ток"]

with pd.ExcelWriter('loads/test.xlsx', engine='xlsxwriter') as writer:
    voltages_table.to_excel(writer, sheet_name='Напряжения', index=False, header=["NetName", "Напряжение, В"])
    #table.to_excel(writer, sheet_name='Потребители', index=False, header=header)
    table_new.to_excel(writer, sheet_name='Потребители', index=False, header=header)

#voltages_table.to_excel("loads/test.xlsx", sheet_name="Напряжения", index=False, header=["NetName", "Напряжение"])
#table.to_excel("loads/test.xlsx", sheet_name="Потребители", index=False)

#print(voltages_table)


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.rule import CellIsRule
header_height = 25
row_height = 20
alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
gray_fill = PatternFill(start_color='E2E2E2',
                        end_color='E2E2E2',
                        fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF',
                        end_color='FFFFFF',
                        fill_type='solid')
red_fill = PatternFill(bgColor='FFC7CE')

rule = CellIsRule(
    operator='lessThan',
    formula=['0'],
    fill=red_fill
)



wb = openpyxl.load_workbook("loads/test.xlsx", data_only=False)


length_pos =  len(table_new[1].unique()) + 5 + 1

columns_n = 100
rows_n = 100

sheet = wb["Напряжения"]
sheet.sheet_properties.tabColor = "1072BA"

for column in ['A', 'B', 'C']:
    sheet.column_dimensions[column].fill = gray_fill
sheet.column_dimensions['C'].width = 500


for cell in sheet[1]:  # первая строка - шапка
    cell.font = Font(bold=True)
    cell.alignment = alignment_style
    cell.border = thin_border
    cell.fill = gray_fill
    

#sheet.row_dimensions[1].height = header_height

for row in sheet.iter_rows(min_row=2):
    i = 0
    for cell in row:
        cell.alignment = alignment_style
        cell.border = thin_border
        if i == 0:
            cell.fill = gray_fill
        i += 1

for row in range(2, 100):
    sheet.row_dimensions[row].height = row_height
    for column in range(1, 3):
        cell = sheet.cell(row=row, column=column)
        cell.alignment = alignment_style
        if row > len(voltages)+1:
            cell.fill = gray_fill

set_column_autowidth(sheet, ['A', 'B'], 1.7)

sheet.conditional_formatting.add(f'B2:B{len(voltages)+1}', rule)

sheet.freeze_panes = 'A2' 
sheet = wb["Потребители"]
sheet.sheet_properties.tabColor = "1272AA"


for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    sheet.column_dimensions[column].fill = gray_fill
sheet.column_dimensions['I'].width = 500

set_column_autowidth(sheet, ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], 1.2)

for cell in sheet[1]:  # первая строка - шапка
    cell.font = Font(bold=True)
    cell.alignment = alignment_style
    cell.border = thin_border
    cell.fill = gray_fill

for row in sheet.iter_rows(min_row=2):
    current_row = row[0].row
    #sheet.row_dimensions[row[0].row].height = row_height
    i = 0
    for cell in row:
        cell.alignment = alignment_style
        cell.border = thin_border
        
        if i == 0:
            cell.fill = gray_fill
        i += 1
        if i == 8:
            cell.value = f'=IF(F{current_row}<>"",F{current_row}*G{current_row},IF(E{current_row}<>"",E{current_row}*G{current_row},IF(D{current_row}<>"",D{current_row}*G{current_row},-1)))'
            #sheet.conditional_formatting.add(cell, rule)

for row in range(2, 100):
    sheet.row_dimensions[row].height = row_height
    for column in range(1, 9):
        cell = sheet.cell(row=row, column=column)
        cell.alignment = alignment_style
        if row > len(table_new)+1:
            cell.fill = gray_fill

rule_formula = FormulaRule(
    formula=['$H2 < 0'],  
    fill=red_fill
)
sheet.conditional_formatting.add(f'F2:F{len(table_new)+1}', rule_formula)
sheet.freeze_panes = 'A2' 
wb.create_sheet(title="Сводная таблица")
sheet = wb["Сводная таблица"]

from openpyxl.worksheet.formula import ArrayFormula

#sheet['A2'] = f"=_xlfn.UNIQUE(Потребители!B2:B50)"
#sheet['B1'] = f"=_xlfn.TRANSPOSE(Напряжения!$A$2:$A$15)"
sheet['A2'] = ArrayFormula(f"A2:A{length_pos}", f'=IFERROR(IF(_xlfn.UNIQUE(Потребители!B2:B{rows_n}) <> "", _xlfn.UNIQUE(Потребители!B2:B{rows_n}), ""), "")')
sheet['B1'] = ArrayFormula(f"B1:{get_column_letter(rows_n)}1", f'=IF(TRANSPOSE(Напряжения!$A$2:$A${rows_n})<>"", TRANSPOSE(Напряжения!$A$2:$A${rows_n}),"")')
#sheet.formula_attributes['B1'] = {'t': 'array', 'ref': f"$B$1:${get_column_letter(16)}$1"}
#sheet['B1'].array = False


#column_letter = get_column_letter(column)

#sheet['B2'] = ArrayFormula(f"B2:{get_column_letter(50)}50", f'=IFERROR(INDEX(Потребители!$H$2:$H$100, MATCH(1, (Потребители!$B$2:$B$100=$A1)*(Потребители!$C$2:$C$100=B$1), 0)), "")')
#value = ArrayFormula(f"B2:{get_column_letter(50)}50", f'=IFERROR(INDEX(Потребители!$H$2:$H$100, MATCH(1, (Потребители!$B$2:$B$100=$A{row})*(Потребители!$C$2:$C$100={column_letter}$1), 0)), "")')
#sheet.cell(row=row, column=column, value=value)

for row in range(2, length_pos+1):
    for column in range(2,columns_n):
        column_letter = get_column_letter(column)
        value = ArrayFormula(f"{column_letter}{row}:{column_letter}{row}", f'=IF(IFERROR(INDEX(Потребители!$H$2:$H$100, MATCH(1, (Потребители!$B$2:$B$100=$A{row})*(Потребители!$C$2:$C$100={column_letter}$1), 0)), "")<>0, IFERROR(INDEX(Потребители!$H$2:$H$100, MATCH(1, (Потребители!$B$2:$B$100=$A{row})*(Потребители!$C$2:$C$100={column_letter}$1), 0)), ""), "")')
        sheet.cell(row=row, column=column, value=value)
        #sheet.cell(row=row, column=column, value=f'=IFERROR(INDEX(Потребители!$H$2:$H$100, MATCH(1, (Потребители!$B$2:$B$100=$A{row})*(Потребители!$C$2:$C$100={column_letter}$1), 0)), "")')


for cell in sheet[1]:  # первая строка - шапка
    cell.font = Font(bold=True)
    cell.alignment = alignment_style
    #cell.border = thin_border
    #cell.fill = gray_fill
    cell.alignment = Alignment(textRotation=45)

sheet.row_dimensions[1].height = 100
for column in range(2,columns_n):
    sheet.column_dimensions[get_column_letter(column)].width = 6

sheet_temp = wb["Потребители"]
sheet.column_dimensions['A'].width = sheet_temp.column_dimensions['B'].width

rule_bb = CellIsRule(
    operator='notEqual',
    formula=['""'],  
    border=thin_border
)

rule_b = FormulaRule(
    formula=['AND(B$1 <> "", $A2 <> "")'],  
    border=thin_border
)
rule_bw = FormulaRule(
    formula=['NOT(AND(B$1 <> "", $A2 <> ""))'],  
    fill = gray_fill
)


for row in sheet.iter_rows(min_row=2):
    current_row = row[0].row
    sheet.row_dimensions[row[0].row].height = row_height
    i = 0
    for cell in row:
        

        cell.alignment = alignment_style
        #cell.border = thin_border
        #if i == 0:
            #cell.fill = gray_fill
        i += 1


sheet.conditional_formatting.add(f'B2:{get_column_letter(columns_n)}{length_pos}', rule_b) 
#sheet.conditional_formatting.add(f'B2:{get_column_letter(50)}50', rule_bw) 
sheet.conditional_formatting.add(f'A1:A{rows_n}', rule_bb) 
sheet.conditional_formatting.add(f'A1:{get_column_letter(columns_n)}1', rule_bb) 
sheet.conditional_formatting.add(f'B2:{get_column_letter(columns_n)}{rows_n}', rule)  

row = length_pos + 2

sheet.row_dimensions[row].height = row_height
cell = sheet.cell(row=row, column=1, value="Напряжение, В")
cell.alignment = alignment_style
cell.border = thin_border

sheet[f'B{row}'] = ArrayFormula(f"B{row}:{get_column_letter(rows_n)}{row}", f'=IF(TRANSPOSE(Напряжения!$B$2:$B${rows_n})<>"", TRANSPOSE(Напряжения!$B$2:$B${rows_n}),"")')

for column in range(2, columns_n):
    cell = sheet.cell(row=row, column=column)
    cell.alignment = alignment_style

row += 1

sheet.row_dimensions[row].height = row_height
cell = sheet.cell(row=row, column=1, value="Суммарный ток, А")
cell.alignment = alignment_style
cell.border = thin_border

for column in range(2, columns_n):
    sheet.cell(row=row, column=column, value= f'=IF(SUM({get_column_letter(column)}2:{get_column_letter(column)}{row-2})>0,SUM({get_column_letter(column)}2:{get_column_letter(column)}{row-2}),"")')
    cell = sheet.cell(row=row, column=column)
    
    cell.alignment = alignment_style
    #cell.border = thin_border

#sheet.conditional_formatting.add(f'B{row}:{get_column_letter(columns_n)}{row}', rule_b)
    
row += 1

sheet.row_dimensions[row].height = row_height
cell = sheet.cell(row=row, column=1, value="Мощность, Вт")
cell.alignment = alignment_style
cell.border = thin_border

for column in range(2, columns_n):
    sheet.cell(row=row, column=column, value= f'=IF(IFERROR({get_column_letter(column)}{row-2}*{get_column_letter(column)}{row-1}, "") > 0, IFERROR({get_column_letter(column)}{row-2}*{get_column_letter(column)}{row-1}, ""), "")')
    cell = sheet.cell(row=row, column=column)
    cell.alignment = alignment_style

sheet.conditional_formatting.add(f'B{row-2}:{get_column_letter(columns_n)}{row}', rule_b)

row += 1
sheet.row_dimensions[row].height = row_height
cell = sheet.cell(row=row, column=1, value="Суммарная мощность, Вт")
cell.alignment = alignment_style
cell.border = thin_border
cell = sheet.cell(row=row, column=2)
cell.alignment = alignment_style
cell.border = thin_border
cell.value = f'=SUM(B{row-1}:{get_column_letter(columns_n)}{row-1})'

sheet.freeze_panes = 'A2' 

wb.save("loads/test.xlsx")



